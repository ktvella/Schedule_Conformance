# -*- coding: utf-8 -*-
"""
Created on Mon May 19 10:37:00 2025
@author: kvella


Script to help automate calculation of schedule conformance in tandem with daily automated
exports from XA. Daily exports should be saved as a CSV file in the format of 
'Monday Sched Conform Wk9.csv'. 


"""

# update these values for each day/week
week = 23
weekday = "Saturday" 


import pandas as pd
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from itertools import zip_longest



def beginning_end_of_week(): 
    """
    Returns
    -------
    beginning_of_week : date
        beginning of current week (sunday)
    end_of_week : date
        end of current week (saturday).
    """
    now = datetime.now()
    now = now.date()
    beginning_of_week = datetime.now().date()-timedelta(days = now.weekday()+1)
    beginning_of_week = pd.to_datetime(beginning_of_week)
    end_of_week = beginning_of_week+timedelta(days = 7)
    end_of_week = pd.to_datetime(end_of_week)
    return beginning_of_week, end_of_week


weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday" ]
def weekday_name_to_num(day_name):
    #Convert weekday name (string) to a number
    for i in range(len(weekdays)): 
        if weekdays[i] == day_name: 
            return i


#find the beginning and end of the current week
beginning_of_week, end_of_week = beginning_end_of_week()
weekday_number = weekday_name_to_num(weekday)    


def parse_date(date_string): 
    """
    function to parse strings to datetime in two possible formats
    """
    for fmt in ('%m/%d/%y', '%m/%d/%Y'): 
        try: 
            return pd.to_datetime(date_string, format = fmt)
        except ValueError: 
            continue
    return pd.NaT

    

def df_cleaning(df): 
    """
    Funtion to clean raw export data
    
    Parameters
    ----------
    df : dataframe
        XA export of scheduled operations.

    Returns
    -------
    df : dataframe
        cleaned dataframe,subsetted to only schedule conformance facilities, 
        renamed columns, columns converted to appropriate format, .
    """
    new_names = {"Department": "Dept", "Item Description": "Description", "Start Date": "Sch start", 
                 "Actual Start Date": "Act start", "Complete Date": "Sch comp", "Actual Completion Date": "Act comp", 
                 "MOP  QTY Remaining": "Qty Rem", "MO Qty Remaining": "Qty Rem MO", "Last Activity Date": "Last activity", "Due date": "Due", 
                "Mach Hrs Remaining": "Mach hrs rem", "Labor Hrs Remaining": "Labor hrs rem"}
    df = df.rename(columns = new_names)             
    sch_conf_facilities = ["MACH51", "MACH52", "MACH53", "MACH54", "MACH55", "MACH56", "MACH58", "MACH48", "MACH49", "MACH50", 
                   "MACH47", "MACH57", "MACH59", "MACH60", "MACH61", "MACH62","MACH63" ,
                   "MACH2", "MACH5", "MACH14", "MACH15", "MACH16", "MACH17", "MACH18", "MACH19", 
                    "MACH20", "MACH99"]
    df = df[df["Facility"].isin(sch_conf_facilities)]
    #df = df[~df["Order"].isin(mos_to_remove)]
    df = df.replace(",", "", regex = True) 
    cols_to_numeric = ["Qty Rem", "Qty Rem MO", "Mach hrs rem", "Labor hrs rem", "Hours Remaining"]  #convert two numeric column to numbers from objects
    df[cols_to_numeric] = df[cols_to_numeric].apply(pd.to_numeric)
    cols_to_date = ["Sch start", "Act start", "Sch comp", "Act comp", "Due", "Last activity"]
    for col in cols_to_date: 
        df[col] = parse_date(df[col])
    return df




#load files for each weekday up to today
def create_by_day_dictionaries(weekday):
    """
    Parameters
    ----------
    weekday : string
        name of current weekday.

    Returns
    -------
    LRP_by_day : dictionary
        dictionary with a key, dataframe of scheduled MOs for each weekday.
    not_sched_by_day : dictionary
        dictionary with a key: dataframe of not scheduled MOs for each weekday.

    """
    LRP_by_day = {}
    not_sched_by_day = {}
    for number in range(weekday+1): 
        #loop through all weekdays so far in week, for each day add df of MOs to dictionary
        #add to either scheduled MOs dict (LRP by day) or not scheduled MOs dict
        weekday_name = weekdays[number]
        df = pd.read_csv(weekday_name +" Sched Conform Wk"+str(week)+".csv")
        df = df_cleaning(df)
        df2 = df.copy()
        df = df[df["Act comp"].isna()]            #select only not complete MOs
        if weekday_name == "Monday":
            #if today is monday, find only scheduled MOs this week by using sch comp date
            mask =  (df["Sch comp"] <= end_of_week)
            df = df.loc[mask]
        df = df[["Order", "Description", "Item", "Qty Rem", "Facility", "Dept", "Mach hrs rem"]] #grab only columns needed 
        df=df[df["Qty Rem"]  > 0]                       #grab only rows where qty remaining is >0
        LRP_by_day[weekday_name] = df
        df2 = df2[~df2["Order"].isin(LRP_by_day["Monday"]["Order"])]   #for not scheduled, select only MOs not in scheduled
        df2 = df2[df2.apply(lambda row: beginning_of_week<row["Last activity"] < end_of_week,  axis = 1)]  #select MOs where last activity is in the current week
        df2 = df2[df2.apply(lambda row: not(row["Act comp"] < pd.to_datetime('today', format = '%Y-%m-%d')),  axis = 1)]
        not_sched_by_day[weekday_name] = df2
    return LRP_by_day, not_sched_by_day


LRP_by_day, not_sched_by_day = create_by_day_dictionaries(weekday_number)

#define department facilities
dept_facilities = {"DeptD": ["MACH51", "MACH52", "MACH53", "MACH54", "MACH55", "MACH57", "MACH58"], 
                   "DeptE": ["MACH48", "MACH49", "MACH50"], 
                   "DeptF": ["MACH47", "MACH56", "MACH59", "MACH60", "MACH61", "MACH63"], 
                   "DeptL": ["MACH62"],
                   "DeptB": ["MACH2", "MACH5", "MACH14", "MACH15", "MACH16", "MACH17", "MACH18", "MACH19", 
                    "MACH20", "MACH99"]}
scheduled_mos = {}
status = {}

def update_status(weekday):
    """
    find a weekdays MOs, split into departments, and update status for each department

    Parameters: weekday (string) 

    returns nothing, but updates existing dataframe 
    """
    
    #helper function to find a weekdays MOs, split into departments, and update status for each department
    todays_scheduled_mos = {}
    for key, value in dept_facilities.items(): 
        source_df = LRP_by_day[weekday]                             #pull data from days export
        df = source_df[source_df["Facility"].isin(value)]     #split into depts by matching facility codes
        df = df.reset_index(drop = True)
        if weekday != "Monday":
            df = pd.merge( scheduled_mos["Monday"][key][["Order", "Description"]], df, on =["Order", "Description"], how = 'inner') #only include MO operations that were in monday's MO list
        todays_scheduled_mos[key] = df
        todays_status = {"Weekday": [weekday], "MO Count": [df["Order"].count()], "Hours":[df["Mach hrs rem"].sum()]}  #create days status
        todays_status = pd.DataFrame(todays_status)
        if weekday == "Monday": 
            status[key] = todays_status
            df_to_export = df[["Order", "Item", "Description", "Mach hrs rem"]]
            with pd.ExcelWriter(key + " Monday Scheduled MOs WK" +str(week)+".xlsx") as writer: 
                df_to_export.to_excel(writer, sheet_name = "Sheet1",  index = False)
            workbook = load_workbook(key + " Monday Scheduled MOs WK" +str(week)+".xlsx")
            fit_column_width(workbook["Sheet1"])
            workbook.save(key + " Monday Scheduled MOs WK" +str(week)+".xlsx")
        else: 
            status[key] = pd.concat([status[key], todays_status], ignore_index = True)  #append tuesday status to existing department df in status dictionary
    scheduled_mos[weekday] = todays_scheduled_mos
    

def generate_reasons(): 
    """
    Loops through each department
    Writes a df with unfinished MOs in that dept to sheet1 in a new spreadsheet
    Add data validation to the next two columns of the spreadsheet

    
    Add the list of reasons and status in Columns A and B on sheet 2 of excel file
    Then adds a status list data validation in row D, reasons list in row E

    calls fit_column_width function to fit columns on both sheets to the length of text contained in the columns

    """
    #define reasons and status
    reasons_list = ["Safety Stop/hold", "Quality hold - NCR", "metals/materials not in stock", "metals/materials not prepped", 
                "material at OSP", "in-stock material found defective", "no compound - outside supplier", "no compound - in-house (M&P)", 
                "mold/tool not available - needs repair", "mold/tool not available - needs cleaning", "insufficient qty of material", 
                "prior work order not complete", "equipment not operational", "equipment under maintenance/PM", 
                "equipment/process not released by Tech/Mfg Eng", "Engineering hold (Design/Product)", "failed bat heat/test", 
                "replaced by expedited work order", "1st pcs failed", "no operator", "documentation error", "waiting on Test Lab", 
                "insufficient time/over scheduled", "hold over from prior week"]
    
    status_list = ["not started", "in progress", "completed"]
    df_reasons = pd.DataFrame(list(zip_longest(reasons_list, status_list)),  columns = ["Reasons", "Status"])
    
    #write dataframe and reasons/status df to excel workbook sheets
    for key, value in status.items(): 
        workbook_name = f"{key} Sch Conf Reasons WK{week}.xlsx"
        with pd.ExcelWriter(workbook_name) as writer:
            df_to_export = scheduled_mos["Friday"][key]
            df_to_export = df_to_export[["Order", "Item", "Mach hrs rem"]]
            df_to_export.insert(3, "Status", "")
            df_to_export.insert(4, "Reason", "")
            df_to_export.insert(5, "Comment", "")
            df_to_export.to_excel(writer, sheet_name = "Sheet1", index = False)
            df_reasons.to_excel(writer, sheet_name ='Sheet2', index = False)
        workbook = load_workbook(workbook_name)
        sheet1 = workbook["Sheet1"] # Sheet where you want the dropdown
        sheet2 = workbook["Sheet2"] # Sheet containing the list
        list_range = "Sheet2!$A$2:$A$25" #location of reasons list
        dv = DataValidation(type="list", formula1=list_range, allow_blank=True)
        sheet1.add_data_validation(dv)
        #add data validation to reasons column E
        for row in range(2, 22): # For rows 2 to 22
             dv.add(sheet1[f"E{row}"])
        list_range2 = "Sheet2!$B$2:$B$4" #location of status list
        dv2 = DataValidation(type = "list", formula1 = list_range2, allow_blank = True)
        #add data validation to status column D
        sheet1.add_data_validation(dv2)
        for row in range(2, 22):
            dv2.add(sheet1[f"D{row}"])
 
     #set column widths to fit text
        fit_column_width(sheet1)
        fit_column_width(sheet2)
 
        workbook.save(workbook_name)  
   

    
    
def fit_column_width(sheet):
    """
    with an existing spreasheet, change column width to fit data in each column
    Parameters: 
        sheet ("string"): name of sheet to make edits to
    """
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width        

def split_by_dept(df):
    """
    Split dataframe into multiple dataframes by department
    
    Parameters: 
    df (dataframe) 
    works with dept_facilities dictionary which has the department as the keys and a list of the corresponding facilities for that dept as the values
    
    Returs a dictionary with the department as the keys, and the df for that department as the value
    
    """
    
    dept_dict = {}
    for key, value in dept_facilities.items(): 
        new_df = df[df["Facility"].isin(value).values]           #split into depts by matching facility codes
        new_df = new_df.reset_index(drop = True)
        dept_dict[key] = new_df
    return dept_dict


#loops through not scheduled by day dictionary, finds first occurence of MO, adds to new df
def setup_not_sched_statuses(not_scheduled_dict): 
    """
    Find the not-scheduled MOs for each department

    Parameters: 
    not_scheduled_dict (dictionary): dictionary where key is day of the week and value is a df of not-scheduled MOs for that day

    returns: dept_dict (dictionary): dictionary where the key is the department and the value is a df of not-scheduled MOs for the week for that
    department, and the progress of those MOs through the week 
    
    """

    for key, value in not_scheduled_dict.items(): 
        if key == "Monday":                          
            # if today is monday, initialize emtpy df
            df = value[["Order", "Description", "Item", "Facility", "Qty Rem MO", "Mach hrs rem"]]
            df.columns = ["MO", "Description", "Item", "Facility", "Initial Qty", "Initial Mach Hrs"]
        else: 
            #if today is not monday, find orders that are new and add to df
            #df_new_today = value[~value["Order"].isin(df["MO"])]
            df_new_today = value
            df_new_today = df_new_today[["Order", "Description", "Item", "Facility", "Qty Rem MO", "Mach hrs rem"]]
            df_new_today.columns = ["MO", "Description", "Item", "Facility", "Initial Qty", "Initial Mach Hrs"]
            df = pd.concat([df, df_new_today], ignore_index = True)
            df.drop_duplicates(subset = ["MO", "Description"], inplace = True)
        
    #find the current day's qty and hrs remaining for MOs in df above
    #add to df as the end qty and hours
    df2 = not_sched_by_day[weekday]                                       #find all orders in the not scheduled df
    df2 = df2[df2["Order"].isin(df["MO"])]                                #select only orders that exist in the newly created df
    df2 = df2[["Order", "Description", "Qty Rem MO", "Mach hrs rem"]]     #select only relevent columns
    df2.columns = ["MO", "Description","End Qty", "End Mach Hrs"]         #rename columns as end values

    
    #combine start and end qty dfs, calculate difference (progress)
    df = pd.merge( df, df2, on = ["MO", "Description"], how = 'left')
    df = df.fillna(0)
    cols_to_numeric = ['Initial Qty', "Initial Mach Hrs", "End Qty", "End Mach Hrs"]  #convert two numeric column to numbers from objects
    df[cols_to_numeric] = df[cols_to_numeric].apply(pd.to_numeric)
    df["Qty Comp"] = df.apply(lambda row: row["Initial Qty"] - row["End Qty"], axis = 1)                  #calculate qty complete so far
    df["Mach Hrs Comp"] = df.apply(lambda row: row["Initial Mach Hrs"] - row["End Mach Hrs"], axis = 1)   #calculate hrs complete so far
    dept_dict = split_by_dept(df)                                                                        #split into dfs by department, store in dictionary 
    return dept_dict
    
def calc_progress(df):
    """
    Calculate progress of MOs and Hrs throughouth the week
    Parameters: 
    df (dataframe): df with columns [Weekday, MO Count, Hours]

    Modifies the df, adds columns [MOs Complete, Hours Complete, %MOs Complete, %Hrs Complete]
    Calulates values from difference between rows
    
    """
    
    #Calculate progress of MOs and Hrs throughouth the week
    if len(df) <= 1: 
        return 
    df["MOs Complete"] = df["MO Count"] - df["MO Count"].shift(-1)   #create a new column of number of MOs completed each day
    df["MOs Complete"] = df["MOs Complete"].cumsum()                 #change that column to cumulative sum of MOs completed
    df["Hours Complete"] = df["Hours"] - df["Hours"].shift(-1)       #create a new column of number of hrs completed each day
    df["Hours Complete"] =df["Hours Complete"].cumsum()              #change that column to a cumulative sum of hrs completed
    monday_mos = df["MO Count"].iloc[0]                              #find mondays # of MOs 
    monday_hrs = df["Hours"].iloc[0]                                 #find mondays # of hours
    df["% MOs Complete"] = round((df["MOs Complete"]/monday_mos)*100, 2)   #create a new column calculating % of MOs completed 
    df["% Hrs Complete"] = round((df["Hours Complete"]/monday_hrs)*100, 2)  #create a new column calculating % of hrs completed
    
for number in range(weekday_number+1):
    day = weekdays[number]
    update_status(day)

not_scheduled_dict = setup_not_sched_statuses(not_sched_by_day)       


def write_to_excel(dictionary, name):
    """
    Parameters
    ----------
    dictionary : dictionary (format key (string): value(df))
        dictionary with a key and correspinding df for each department.
    name : string
        name of file you want to save

    writes all dfs in a dictionary to an excel file
    name sheets by the keys
    fit column width to width of text

    """
    with pd.ExcelWriter(name) as writer: 
        for key, value in dictionary.items(): 
            dictionary[key].to_excel(writer,sheet_name = key, index = False)
    workbook = load_workbook(name)
    for key, value in dictionary.items(): 
        sheet = workbook[key]
        fit_column_width(sheet)
    workbook.save(name)


"""

Exporting dataframes to excel

Not Scheduled MOs file: workbook with a sheet for each department of not scheduled MOs and progress 
Sch Conf Status file: workbook with a sheet for each department of scheduled MOs and progress
APU Sch Conf Reasons file: a file for each APU, exported at end of week, with not completed scheduled MOs
    for apu managers to fill with reasons not complete

"""

#save not scheduled mos to file with sheet for each department
write_to_excel(not_scheduled_dict ,"Not Scheduled MOs WK" +str(week) + ".xlsx")

#calculate schedule conformance progress 
def run_status_calcs(): 
    for key, value in status.items(): 
        calc_progress(status[key])

run_status_calcs()

#write status dfs to workbook
write_to_excel(status, "Sch Conf Status WK" + str(week) +".xlsx")


#on saturday, create final statuses and reasons spreadsheet
if weekday_number >= 5:  
   generate_reasons()   

 
    
"""
Generating Graphs and Paretos

each day a plot is generated to show the progress of MOs for each APU, as well as machine hours
at the end of the week, the pareto of reasons for each department is generated

"""

def generate_2_plot():
    """
    Creates plots with 2 subplots, one for % MOs completed and one for % Hrs completed
    
    Returns
    -------
    None.
    Outputs plot of variables
    
    """
    #target = [20, 40, 60, 80, 100]
    #legend_labels = list(status.keys())
    #legend_labels.append("Target")

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize = (7,8))
    fig.tight_layout()
    plt.subplots_adjust(hspace = .2)
    generate_subplots("% MOs Complete", ax1)
    generate_subplots("% Hrs Complete", ax2)
    plt.savefig( "Status Week " +str(week), bbox_inches = "tight")


def generate_subplots(variable, axis): 
    """
    plots subplots on given axis 
    
    Parameters
    ----------
    variable : string
        variable to plot, either '% MOs Complete' or '% Labor Complete' .
    axis: axis to plot variable on

    
    """
    
    target = [20, 40, 60, 80, 100]
    legend_labels = list(status.keys())
    legend_labels.append("Target")
    
    if variable == "% MOs Complete": 
        string1 = "MO Status"
        string2 = "MOs"
        y_loc = 190
    else: 
        string1 = "Labor Status"
        string2 = "Hours"
        y_loc = 70
    
    if weekday_number >0: 
        x = list(range(weekday_number+1))
        markers = ["o", "x", "s", "d", "s"]
        colors = ["red", "blue", "green", "cyan", "orange"]
        num = 0
        for key, value in status.items(): 
            axis.plot(x, status[key][variable], alpha = .7, color = colors[num], marker = markers[num])
            num+=1
        axis.plot(x[:weekday_number], target[:weekday_number], alpha =.7, color = 'navy', linestyle = '--')
        axis.legend(legend_labels)
        axis.set_title(string1)
        #axis.set_xlabel("Day of the Week")
        axis.set_xticks(x, weekdays[:weekday_number+1])
        axis.set_ylabel("% Complete")
        axis.grid(True)
        axis.set_ylim(0, 100)
        text1 = f"DeptD % of {string2} Complete: {status["DeptD"].at[weekday_number-1, variable]}% "
        text2 = f"DeptE % of {string2} Complete: {status["DeptE"].at[weekday_number-1, variable]}%"
        text3 = f"DeptF % of {string2} Complete: {status["DeptF"].at[weekday_number-1, variable]}%"
        text4 = f"DeptL % of {string2} Complete: {status["DeptL"].at[weekday_number-1, variable]}%"
        text5 = f"DeptB % of {string2} Complete: {status["DeptB"].at[weekday_number-1, variable]}%"
        plt.text( weekday_number+.05, y_loc, text1 + "\n" + text2 +"\n" +text3 + "\n" +text4 + "\n" +text5)


generate_2_plot()
  
    

    

